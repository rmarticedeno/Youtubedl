import youtube_dl
from YVideo import YVideo
from openpyxl import Workbook


def Get_Videos(url):
    '''
    Get all youtube videos from given url
    Output: YVideo List (see YVideo.py for more info)
    '''

    ydl = youtube_dl.YoutubeDL({'outtmpl': '%(id)s.%(ext)s'})

    with ydl:
        result = ydl.extract_info(
            url,
            download=False
        )

    videos = []

    if 'entries' in result:
        # playlistm channel or list of videos
        for entry in result['entries']:
            videos.append( 
                YVideo(
                    entry['title'],
                    entry['id']
                    )
                )
    else:
        # Just a video
        videos.append( 
                YVideo(
                    result['title'],
                    result['id']
                    )
                )

    return videos

def Save_to_Excel(vlist, name):
    '''
    Save YVideos List to name.xlsx file
    '''

    wb = Workbook()
    ws = wb.active

    for i,video in enumerate(vlist):
        ws.cell(row=i+1, column=1, value=video.title)
        ws.cell(row=i+1, column=2, value=video.url)
    
    wb.save(f'{name}.xlsx')


def Get_To_Excel(url, name):
    '''
    Extract Yvideos from url to name.xlsx file
    '''

    vids = Get_Videos(url)
    Save_to_Excel(vids, name)

